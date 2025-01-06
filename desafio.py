#percorrer toda a base de dados e verficar se o setor possui um aba
from openpyxl import load_workbook

def criar_aba(segmento, arquivo_fiis):
    if segmento not in arquivo_fiis.sheetnames:
        arquivo_fiis.create_sheet(segmento)

def transferir_informacoes_aba(aba_origem, aba_destino, linha_origem):
    linha_destino = aba_destino.max_row +1
    for coluna in range(1, 4):
        celula_origem = aba_origem.cell(row=linha_origem, column=coluna)
        celula_destino = aba_destino.cell(row=linha_destino, column=coluna)
        celula_destino.value = celula_origem.value 

arquivo_fiis = load_workbook("Dados/Basededados.xlsx")

print(arquivo_fiis.sheetnames)

aba_Planilha1 = arquivo_fiis["Planilha1"]

ultima_linha = aba_Planilha1.max_row
print(ultima_linha)
#ou
#for linha in range(1, 373):
    #print(linha)

for linha in range(2, ultima_linha + 1):
    segmento = aba_Planilha1.cell(row=linha, column=2).value
    #ou bairro = aba_Planilha1[f"B{linha}"].value
    if not segmento:
        break
    # Criar uma aba para Bairro
    criar_aba(segmento, arquivo_fiis)

    # transferir as informações para aba
    aba_destino = arquivo_fiis["Planilha1"]
    transferir_informacoes_aba(aba_Planilha1, aba_destino, linha)
   



    arquivo_fiis.save("base2.xlsx")