from openpyxl import load_workbook

arquivo = load_workbook('Dados/Basededados.xlsx')

# ver as abas
print(arquivo.sheetnames)

#pegar a aba ativa
aba_atual = arquivo.active
print(aba_atual)

# selecionar aba específica
aba_alunos = arquivo["Planilha1"]
print(aba_alunos)

# selecionar celulas
valor_a1 = aba_alunos["A1"].value
valor_b1 = aba_alunos.cell(row=1, column=2).value
print(valor_b1)

aba_alunos.cell(row=1, column=2).value = "Setor"

arquivo.save("base2.xlsx")

#ultima linha e ultima coluna
print(aba_alunos.max_row)
print(aba_alunos.max_column)
